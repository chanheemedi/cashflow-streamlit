import streamlit as st
import pandas as pd
import numpy as np
import re
import hashlib
from io import BytesIO
from datetime import datetime, timedelta, date
import calendar
import math
import openpyxl

# Google Sheets
import gspread
from google.oauth2.service_account import Credentials

# =========================================================
# Config
# =========================================================
EXCLUDED_ACCOUNT_NAMES = {"신한_에셀", "하나_꾸러기건식"}  # 병원 범위에서 제외
INTERNAL_WINDOW = timedelta(hours=2)

DEFAULT_DRAWDOWN_KEYWORDS = ["메디컬네트워크론"]
PRINCIPAL_KEYWORDS = ["메디칼론원금"]
INTEREST_KEYWORDS = ["메디칼론이자"]

DOW_KO = ["월", "화", "수", "목", "금", "토", "일"]

st.set_page_config(page_title="현금흐름 MVP", layout="wide")

# =========================================================
# DB (Google Sheets) helpers
# =========================================================
def _normalize_sheet_id(x: str) -> str:
    s = str(x).strip()
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", s)
    if m:
        return m.group(1)
    s = s.split("/edit")[0]
    return s

def _db_enabled() -> bool:
    try:
        return ("gcp_service_account" in st.secrets) and ("app" in st.secrets) and ("sheet_id" in st.secrets["app"])
    except Exception:
        return False

DB_ENABLED = _db_enabled()

@st.cache_resource
def gs_client():
    info = dict(st.secrets["gcp_service_account"])
    # secrets에 \n 텍스트로 들어와도 실제 개행으로 변환
    if "private_key" in info and isinstance(info["private_key"], str):
        info["private_key"] = info["private_key"].replace("\\n", "\n")

    creds = Credentials.from_service_account_info(
        info,
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ],
    )
    return gspread.authorize(creds)

def ws(name: str):
    sheet_id = _normalize_sheet_id(st.secrets["app"]["sheet_id"])
    sh = gs_client().open_by_key(sheet_id)
    return sh.worksheet(name)

def read_df(sheet_name: str) -> pd.DataFrame:
    if not DB_ENABLED:
        return pd.DataFrame()
    w = ws(sheet_name)
    values = w.get_all_values()
    if not values:
        return pd.DataFrame()
    header = values[0]
    rows = values[1:]
    df = pd.DataFrame(rows, columns=header)

    # 완전 빈 줄 제거
    df = df.replace("", np.nan).dropna(how="all").fillna("")
    return df

def overwrite_df(sheet_name: str, df: pd.DataFrame):
    if not DB_ENABLED:
        return
    w = ws(sheet_name)
    w.clear()
    if df is None or df.empty:
        # 헤더만이라도 유지하려면 필요시 여기에서 처리 가능
        return
    w.update([df.columns.tolist()] + df.astype(str).fillna("").values.tolist())

# =========================================================
# internal_transfers DB helpers
# =========================================================
IT_SHEET = "internal_transfers"
IT_COLS = ["out_tx_id", "in_tx_id", "status", "updated_at", "note"]
IT_STATUSES = ["AUTO", "CONFIRMED", "REJECTED"]  # AUTO=미결(저장안함)
DEFAULT_NEW_CAND_STATUS = "CONFIRMED"

def load_it_db() -> pd.DataFrame:
    """Google Sheets internal_transfers 로드(중복키는 최신 updated_at 1건만 유지)."""
    if not DB_ENABLED:
        return pd.DataFrame(columns=IT_COLS)

    df = read_df(IT_SHEET)
    if df is None or df.empty:
        return pd.DataFrame(columns=IT_COLS)

    # 컬럼 보정
    for c in IT_COLS:
        if c not in df.columns:
            df[c] = ""

    df = df[IT_COLS].copy()
    df = df.replace("", np.nan).dropna(subset=["out_tx_id", "in_tx_id"], how="any").fillna("")

    df["status"] = df["status"].astype(str).str.upper().replace({"TRUE":"CONFIRMED","FALSE":"REJECTED"})
    df.loc[~df["status"].isin(["CONFIRMED","REJECTED"]), "status"] = ""

    # 최신 1건 유지(문자열 시간 정렬로 충분)
    df = df.sort_values("updated_at", na_position="last")
    df = df.drop_duplicates(subset=["out_tx_id", "in_tx_id"], keep="last")
    return df

def it_map_from_df(df: pd.DataFrame) -> dict:
    """{(out,in): {'status':..., 'note':...}}"""
    m = {}
    if df is None or df.empty:
        return m
    for _, r in df.iterrows():
        k = (str(r["out_tx_id"]), str(r["in_tx_id"]))
        stt = str(r.get("status","")).upper()
        note = str(r.get("note",""))
        if stt in ["CONFIRMED","REJECTED"]:
            m[k] = {"status": stt, "note": note}
    return m

def upsert_it_db(decisions: dict) -> None:
    """
    decisions: {(out,in): {'status': 'CONFIRMED'|'REJECTED', 'note': '...'}}
    DB는 (out,in) 키로 upsert.
    """
    if not DB_ENABLED:
        return

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur = load_it_db()

    # 현재 DB를 map으로
    cur_map = {}
    if not cur.empty:
        for _, r in cur.iterrows():
            cur_map[(str(r["out_tx_id"]), str(r["in_tx_id"]))] = {
                "status": str(r.get("status","")).upper(),
                "updated_at": str(r.get("updated_at","")),
                "note": str(r.get("note","")),
            }

    # 업데이트 반영
    for k, v in decisions.items():
        stt = str(v.get("status","")).upper()
        note = str(v.get("note",""))
        if stt not in ["CONFIRMED","REJECTED"]:
            continue
        cur_map[(str(k[0]), str(k[1]))] = {
            "status": stt,
            "updated_at": now,
            "note": note,
        }

    # 다시 DF로 저장
    out_rows = []
    for (out_id, in_id), vv in cur_map.items():
        out_rows.append({
            "out_tx_id": out_id,
            "in_tx_id": in_id,
            "status": vv.get("status",""),
            "updated_at": vv.get("updated_at",""),
            "note": vv.get("note",""),
        })
    out_df = pd.DataFrame(out_rows, columns=IT_COLS).sort_values("updated_at", na_position="last")
    overwrite_df(IT_SHEET, out_df)

def sync_confirmed_pairs_from_it_map():
    """it_status_map 기반으로 confirmed_pairs(집계 제외 대상) 갱신"""
    it_map = st.session_state.get("it_status_map", {})
    st.session_state.confirmed_pairs = {k for k,v in it_map.items() if v.get("status")=="CONFIRMED"}
    st.session_state.rejected_pairs = {k for k,v in it_map.items() if v.get("status")=="REJECTED"}


# =========================================================
# Display helpers
# =========================================================
def fmt_int(x):
    if x is None:
        return ""
    try:
        if isinstance(x, (np.integer, int)):
            return f"{int(x):,}"
        if isinstance(x, (np.floating, float)):
            if math.isnan(x):
                return ""
            return f"{int(round(x, 0)):,}"
        if isinstance(x, str) and x.strip() == "":
            return ""
        return f"{int(float(x)):,}"
    except Exception:
        return str(x)

def parse_korean_dt_str(s):
    if pd.isna(s) or s is None:
        return pd.NaT
    ss = str(s).replace("오전", "AM").replace("오후", "PM")
    ss = re.sub(r"\s+", " ", ss).strip()
    return pd.to_datetime(ss, format="%Y/%m/%d %p %I:%M:%S", errors="coerce")

def parse_korean_ampm_series(s: pd.Series) -> pd.Series:
    def _one(x):
        if pd.isna(x):
            return pd.NaT
        x = str(x).replace("오전", "AM").replace("오후", "PM")
        x = re.sub(r"\s+", " ", x).strip()
        return pd.to_datetime(x, format="%Y/%m/%d %p %I:%M:%S", errors="coerce")
    return s.apply(_one)

def month_range(year: int, month: int):
    first = date(year, month, 1)
    last = date(year, month, calendar.monthrange(year, month)[1])
    return first, last

def make_tx_id(posted_at: pd.Timestamp, account_name: str, direction: str, amount: int, counterparty: str | None) -> str:
    base = f"{posted_at.isoformat()}|{account_name}|{direction}|{amount}|{counterparty or ''}"
    return hashlib.sha256(base.encode("utf-8")).hexdigest()

# =========================================================
# Detect file types
# =========================================================
@st.cache_data(show_spinner=False)
def get_sheetnames(file_bytes: bytes) -> list[str]:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    return wb.sheetnames

def detect_template_workbook(file_bytes: bytes) -> bool:
    s = set(get_sheetnames(file_bytes))
    return ("이카운트 DB" in s) and ("월수입지출" in s)

def detect_balance_package(file_bytes: bytes) -> bool:
    s = set(get_sheetnames(file_bytes))
    return ("계좌별 잔액" in s) and ("메디칼론한도여유액" in s) and ("rawdata" in s)

# =========================================================
# Parsers
# =========================================================
@st.cache_data(show_spinner=False)
def parse_template_cashflow(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(BytesIO(file_bytes), sheet_name="이카운트 DB", engine="openpyxl", header=1)

    need = ["입/출금일자", "입/출", "변환금액", "계좌명", "입금처(출금처)", "원화잔액", "거래처코드"]
    for c in need:
        if c not in df.columns:
            raise ValueError(f"[템플릿] 컬럼 누락: {c} / 현재컬럼: {list(df.columns)}")

    df["posted_at"] = pd.to_datetime(df["입/출금일자"], errors="coerce")
    df = df[df["posted_at"].notna()].copy()

    df["amount"] = pd.to_numeric(df["변환금액"], errors="coerce")
    df = df[df["amount"].notna()].copy()
    df["amount"] = df["amount"].astype(int)

    df["direction"] = np.where(df["입/출"].astype(str).str.startswith("출금"), "OUT", "IN")

    df["account_name"] = df["계좌명"].astype(str)
    df["counterparty"] = df["입금처(출금처)"].astype(str)

    df["is_excluded_account"] = df["account_name"].isin(EXCLUDED_ACCOUNT_NAMES)
    df["balance"] = pd.to_numeric(df["원화잔액"], errors="coerce")
    df["biz_date"] = df["posted_at"].dt.date

    df["is_principal"] = (df["counterparty"].isin(PRINCIPAL_KEYWORDS)) & (df["direction"] == "OUT")
    df["is_interest"] = (df["counterparty"].isin(INTEREST_KEYWORDS)) & (df["direction"] == "OUT")
    df["is_drawdown"] = False

    df["subtype"] = np.where(df["입/출"].astype(str).str.contains("청구"), "CLAIM", "NORMAL")

    df["is_internal_auto"] = df["거래처코드"].astype(str).eq("자금이동")

    df["tx_id"] = df.apply(
        lambda r: make_tx_id(r["posted_at"], r["account_name"], r["direction"], int(r["amount"]), r.get("counterparty")),
        axis=1
    )
    df["source"] = "TEMPLATE"

    return df[[
        "tx_id","posted_at","biz_date","account_name","direction","subtype","amount",
        "counterparty","balance","is_excluded_account","is_internal_auto",
        "is_principal","is_interest","is_drawdown","source"
    ]].copy()

@st.cache_data(show_spinner=False)
def read_ecount_export(file_bytes: bytes) -> pd.DataFrame:
    raw = pd.read_excel(BytesIO(file_bytes), engine="openpyxl", header=None).dropna(how="all")
    header_row_candidates = raw.index[
        raw.iloc[:, 0].astype(str).str.contains("입/출금일자", na=False)
    ]
    if len(header_row_candidates) == 0:
        raise ValueError(f"[ECOUNT] 헤더 행을 찾지 못했습니다. 첫 열 샘플: {raw.iloc[:5,0].tolist()}")
    header_row = int(header_row_candidates[0])

    df = pd.read_excel(BytesIO(file_bytes), engine="openpyxl", header=header_row)
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
    df = df.dropna(axis=1, how="all").dropna(how="all")

    need = ["입/출금일자","구분","계좌명","금액"]
    missing = [c for c in need if c not in df.columns]
    if missing:
        raise ValueError(f"[ECOUNT] 필수 컬럼 누락: {missing} / 현재컬럼: {list(df.columns)}")

    if df["입/출금일자"].dtype == object:
        dt_series = parse_korean_ampm_series(df["입/출금일자"])
    else:
        dt_series = pd.to_datetime(df["입/출금일자"], errors="coerce")

    df["posted_at"] = dt_series
    df = df[df["posted_at"].notna()].copy()

    df["direction"] = df["구분"].astype(str).map({"입금":"IN","출금":"OUT"})
    df = df[df["direction"].isin(["IN","OUT"])].copy()

    df["amount"] = pd.to_numeric(df["금액"], errors="coerce").fillna(0).astype(int)
    df = df[df["amount"] > 0].copy()

    df["account_name"] = df["계좌명"].astype(str)
    df["counterparty"] = df["입금처(출금처)"].astype(str) if "입금처(출금처)" in df.columns else ""
    df["balance"] = pd.to_numeric(df["원화잔액"], errors="coerce") if "원화잔액" in df.columns else np.nan

    df["is_excluded_account"] = df["account_name"].isin(EXCLUDED_ACCOUNT_NAMES)

    df["is_principal"] = (df["counterparty"].isin(PRINCIPAL_KEYWORDS)) & (df["direction"] == "OUT")
    df["is_interest"] = (df["counterparty"].isin(INTEREST_KEYWORDS)) & (df["direction"] == "OUT")
    df["is_drawdown"] = False

    df["subtype"] = np.where(
        (df["direction"] == "IN") & (df["counterparty"].astype(str).str.contains("채권", na=False)),
        "CLAIM",
        "NORMAL"
    )

    df["is_internal_auto"] = False
    df["biz_date"] = df["posted_at"].dt.date

    df["tx_id"] = df.apply(
        lambda r: make_tx_id(r["posted_at"], r["account_name"], r["direction"], int(r["amount"]), r.get("counterparty")),
        axis=1
    )
    df["source"] = "ECOUNT_EXPORT"

    return df[[
        "tx_id","posted_at","biz_date","account_name","direction","subtype","amount",
        "counterparty","balance","is_excluded_account","is_internal_auto",
        "is_principal","is_interest","is_drawdown","source"
    ]].copy()

@st.cache_data(show_spinner=False)
def parse_balance_package(file_bytes: bytes) -> tuple[pd.DataFrame, dict]:
    # ---- rawdata sheet
    raw = pd.read_excel(BytesIO(file_bytes), sheet_name="rawdata", engine="openpyxl", header=None).dropna(how="all")
    header_row_candidates = raw.index[
        raw.iloc[:, 0].astype(str).str.contains("입/출금일자", na=False)
    ]
    if len(header_row_candidates) == 0:
        raise ValueError("[balance_package] rawdata 헤더를 찾지 못했습니다.")
    h = int(header_row_candidates[0])
    cols = raw.iloc[h].tolist()
    df = raw.iloc[h+1:].copy()
    df.columns = cols
    df = df.dropna(how="all")

    df["posted_at"] = parse_korean_ampm_series(df["입/출금일자"])
    df = df[df["posted_at"].notna()].copy()

    df["direction"] = df["구분"].astype(str).map({"입금":"IN","출금":"OUT"})
    df = df[df["direction"].isin(["IN","OUT"])].copy()

    df["amount"] = pd.to_numeric(df["금액"], errors="coerce").fillna(0).astype(int)
    df = df[df["amount"] > 0].copy()

    df["account_name"] = df["계좌명"].astype(str)
    df["counterparty"] = df["입금처(출금처)"].astype(str)
    df["balance"] = pd.to_numeric(df["원화잔액"], errors="coerce")

    df["is_excluded_account"] = df["account_name"].isin(EXCLUDED_ACCOUNT_NAMES)

    df["subtype"] = np.where(
        (df["direction"] == "IN") & (df["counterparty"].astype(str).str.contains("채권", na=False)),
        "CLAIM",
        "NORMAL"
    )

    df["is_principal"] = (df["counterparty"].isin(PRINCIPAL_KEYWORDS)) & (df["direction"] == "OUT")
    df["is_interest"] = (df["counterparty"].isin(INTEREST_KEYWORDS)) & (df["direction"] == "OUT")
    df["is_drawdown"] = False

    df["is_internal_auto"] = False
    df["biz_date"] = df["posted_at"].dt.date

    df["tx_id"] = df.apply(
        lambda r: make_tx_id(r["posted_at"], r["account_name"], r["direction"], int(r["amount"]), r.get("counterparty")),
        axis=1
    )
    df["source"] = "BALANCE_PACKAGE"

    # ---- account snapshot
    acc = pd.read_excel(BytesIO(file_bytes), sheet_name="계좌별 잔액", engine="openpyxl")
    acc["snap_time"] = acc["최종성공일시"].apply(parse_korean_dt_str)
    cash_time = acc["snap_time"].max()

    acc["통장잔액_num"] = pd.to_numeric(acc["통장잔액"], errors="coerce")
    bank_bal = dict(zip(acc["계좌명"].astype(str), acc["통장잔액_num"]))

    # ---- loan snapshot
    loan = pd.read_excel(BytesIO(file_bytes), sheet_name="메디칼론한도여유액", engine="openpyxl")
    loan_time = pd.to_datetime(loan.loc[0, "조회시점"])
    loan_avail = int(pd.to_numeric(loan.loc[0, "한도여유액"], errors="coerce"))

    anchor = {
        "cash_time": cash_time,
        "bank_balances": bank_bal,
        "loan_time": loan_time,
        "loan_avail": loan_avail,
    }
    return df[[
        "tx_id","posted_at","biz_date","account_name","direction","subtype","amount",
        "counterparty","balance","is_excluded_account","is_internal_auto",
        "is_principal","is_interest","is_drawdown","source"
    ]].copy(), anchor

def parse_any_excel(file_bytes: bytes) -> tuple[pd.DataFrame, dict | None]:
    if detect_balance_package(file_bytes):
        df, anchor = parse_balance_package(file_bytes)
        return df, anchor
    if detect_template_workbook(file_bytes):
        return parse_template_cashflow(file_bytes), None
    return read_ecount_export(file_bytes), None

# =========================================================
# Internal transfer matching
# =========================================================
def build_internal_candidates(tx: pd.DataFrame) -> pd.DataFrame:
    if tx is None or tx.empty:
        return pd.DataFrame()

    base = tx[
        (~tx["is_excluded_account"]) &
        (~tx["is_principal"]) &
        (~tx["is_interest"])
    ].copy()

    outs = base[base["direction"] == "OUT"].sort_values("posted_at")
    ins  = base[base["direction"] == "IN"].sort_values("posted_at")

    ins_by_amount = {}
    for _, r in ins.iterrows():
        ins_by_amount.setdefault(int(r["amount"]), []).append(r)

    used_in = set()
    rows = []
    for _, o in outs.iterrows():
        pool = ins_by_amount.get(int(o["amount"]), [])
        best = None
        best_diff = None
        for i in pool:
            if i["tx_id"] in used_in:
                continue
            if i["account_name"] == o["account_name"]:
                continue
            diff = abs(i["posted_at"] - o["posted_at"])
            if diff <= INTERNAL_WINDOW:
                if best is None or diff < best_diff:
                    best, best_diff = i, diff
        if best is not None:
            used_in.add(best["tx_id"])
            rows.append({
                "out_tx_id": o["tx_id"],
                "in_tx_id": best["tx_id"],
                "amount": int(o["amount"]),
                "time_diff_seconds": int(best_diff.total_seconds()),
                "out_time": o["posted_at"],
                "out_account": o["account_name"],
                "in_time": best["posted_at"],
                "in_account": best["account_name"],
                "out_counterparty": o.get("counterparty",""),
                "in_counterparty": best.get("counterparty",""),
            })
    return pd.DataFrame(rows)

# =========================================================
# Anchor & balance math
# =========================================================
def balance_at_time_for_account(g: pd.DataFrame, t: pd.Timestamp):
    g = g[g["posted_at"] <= t].sort_values("posted_at")
    cur = None
    for _, r in g.iterrows():
        b = r["balance"]
        amt = int(r["amount"])
        if pd.notna(b):
            cur = float(b)
        else:
            if cur is not None:
                if r["direction"] == "IN":
                    cur += amt
                else:
                    cur -= amt
    return cur

def compute_anchor_cash_from_bank_and_tx(tx: pd.DataFrame, anchor: dict) -> tuple[float, list[str]]:
    cash_time = anchor["cash_time"]
    bank_bal = anchor["bank_balances"]

    hospital_accounts = [a for a in bank_bal.keys() if a and (a not in EXCLUDED_ACCOUNT_NAMES)]
    missing = []
    total = 0.0

    for acct in hospital_accounts:
        v = bank_bal.get(acct)
        if v is None or (isinstance(v, float) and math.isnan(v)):
            g = tx[tx["account_name"] == acct] if tx is not None else pd.DataFrame()
            inferred = balance_at_time_for_account(g, cash_time) if len(g) else None
            if inferred is None:
                missing.append(acct)
                inferred = 0.0
            total += float(inferred)
        else:
            total += float(v)

    return total, missing

def cash_balance_from_anchor(tx: pd.DataFrame, anchor_time: pd.Timestamp, anchor_cash: float, target_time: pd.Timestamp,
                            confirmed_pairs: set[tuple[str,str]]) -> float:
    if tx is None or tx.empty:
        return float(anchor_cash)

    confirmed_out = {p[0] for p in confirmed_pairs}
    confirmed_in = {p[1] for p in confirmed_pairs}
    excluded_ids = confirmed_out.union(confirmed_in)

    base = tx[~tx["is_excluded_account"]].copy()
    base = base[~base["tx_id"].isin(excluded_ids)]

    mask = (base["posted_at"] > target_time) & (base["posted_at"] <= anchor_time)
    seg = base.loc[mask, ["direction","amount"]].copy()
    seg["cash_delta"] = np.where(seg["direction"] == "IN", seg["amount"], -seg["amount"])
    net = float(seg["cash_delta"].sum())
    return float(anchor_cash) - net

def loan_avail_from_anchor(tx: pd.DataFrame, anchor_time: pd.Timestamp, anchor_loan_avail: float, target_time: pd.Timestamp,
                           drawdown_keywords: list[str]) -> float:
    if tx is None or tx.empty:
        return float(anchor_loan_avail)

    base = tx[~tx["is_excluded_account"]].copy()
    mask = (base["posted_at"] > target_time) & (base["posted_at"] <= anchor_time)
    seg = base.loc[mask, ["direction","amount","counterparty","is_principal"]].copy()

    seg["loan_delta"] = 0
    seg.loc[seg["is_principal"] == True, "loan_delta"] = seg.loc[seg["is_principal"] == True, "amount"]

    if drawdown_keywords:
        pat = "|".join([re.escape(k) for k in drawdown_keywords if k.strip()])
        if pat:
            is_dd = (seg["direction"] == "IN") & (seg["counterparty"].astype(str).str.contains(pat, na=False))
            is_dd = is_dd & (~seg["counterparty"].astype(str).str.contains("원금|이자", na=False))
            seg.loc[is_dd, "loan_delta"] = -seg.loc[is_dd, "amount"]

    net = float(seg["loan_delta"].sum())
    return float(anchor_loan_avail) - net

def daily_aggregations(tx: pd.DataFrame, confirmed_pairs: set[tuple[str,str]], drawdown_keywords: list[str]) -> pd.DataFrame:
    if tx is None or tx.empty:
        return pd.DataFrame(columns=[
            "biz_date","actual_in_claim","actual_in","actual_out_oper","principal","interest",
            "drawdown","loan_delta","total_in","total_out_cash","cash_net"
        ])

    confirmed_out = {p[0] for p in confirmed_pairs}
    confirmed_in = {p[1] for p in confirmed_pairs}
    excluded_ids = confirmed_out.union(confirmed_in)

    base = tx[~tx["is_excluded_account"]].copy()
    base = base[~base["tx_id"].isin(excluded_ids)].copy()

    dd = np.zeros(len(base), dtype=bool)
    if drawdown_keywords:
        pat = "|".join([re.escape(k) for k in drawdown_keywords if k.strip()])
        if pat:
            dd = (base["direction"] == "IN") & (base["counterparty"].astype(str).str.contains(pat, na=False))
            dd = dd & (~base["counterparty"].astype(str).str.contains("원금|이자", na=False))
    base["is_drawdown_calc"] = dd

    base["in_claim"] = np.where((base["direction"]=="IN") & (base["subtype"]=="CLAIM"), base["amount"], 0)
    base["in_normal"] = np.where((base["direction"]=="IN") & (base["subtype"]!="CLAIM"), base["amount"], 0)

    base["out_oper"] = np.where((base["direction"]=="OUT") & (~base["is_principal"]) & (~base["is_interest"]), base["amount"], 0)
    base["out_principal"] = np.where(base["is_principal"], base["amount"], 0)
    base["out_interest"] = np.where(base["is_interest"], base["amount"], 0)

    base["drawdown_in"] = np.where(base["is_drawdown_calc"], base["amount"], 0)

    base["loan_delta"] = 0
    base.loc[base["is_principal"], "loan_delta"] = base.loc[base["is_principal"], "amount"]
    base.loc[base["is_drawdown_calc"], "loan_delta"] = -base.loc[base["is_drawdown_calc"], "amount"]

    g = base.groupby("biz_date").agg(
        actual_in_claim=("in_claim","sum"),
        actual_in=("in_normal","sum"),
        actual_out_oper=("out_oper","sum"),
        principal=("out_principal","sum"),
        interest=("out_interest","sum"),
        drawdown=("drawdown_in","sum"),
        loan_delta=("loan_delta","sum"),
    ).reset_index()

    g["total_in"] = g["actual_in_claim"] + g["actual_in"]
    g["total_out_cash"] = g["actual_out_oper"] + g["principal"] + g["interest"]
    g["cash_net"] = g["total_in"] - g["total_out_cash"]
    return g

def build_month_table(tx: pd.DataFrame, plan_df: pd.DataFrame, year: int, month: int,
                      anchor_info: dict | None,
                      confirmed_pairs: set[tuple[str,str]],
                      drawdown_keywords: list[str]) -> pd.DataFrame:
    first, last = month_range(year, month)
    days = pd.date_range(first, last, freq="D")

    # 계획(예정) -> 일자 합산 map
    plan_map = {}
    if plan_df is not None and len(plan_df) > 0:
        p = plan_df.copy()
        p["biz_date"] = pd.to_datetime(p["date"], errors="coerce").dt.date
        p = p[p["biz_date"].notna()].copy()
        p["amount"] = pd.to_numeric(p["amount"], errors="coerce").fillna(0).astype(int)
        p["direction"] = p["direction"].astype(str).str.upper().replace({"입금":"IN","출금":"OUT"})
        p["plan_in"] = np.where(p["direction"]=="IN", p["amount"], 0)
        p["plan_out"] = np.where(p["direction"]=="OUT", p["amount"], 0)
        plan_map = p.groupby("biz_date").agg(plan_in=("plan_in","sum"), plan_out=("plan_out","sum")).to_dict("index")

    # 실제 집계
    daily = daily_aggregations(tx, confirmed_pairs, drawdown_keywords)
    daily_map = daily.set_index("biz_date").to_dict("index") if len(daily) else {}

    rows = []

    # 앵커 기반 시작값
    if anchor_info is not None:
        anchor_time = anchor_info["anchor_time"]
        anchor_cash = anchor_info["anchor_cash"]
        anchor_loan = anchor_info["anchor_loan_avail"]

        month_start_ts = pd.Timestamp(datetime.combine(first, datetime.min.time()))
        cash_start = cash_balance_from_anchor(tx, anchor_time, anchor_cash, month_start_ts, confirmed_pairs)
        loan_start = loan_avail_from_anchor(tx, anchor_time, anchor_loan, month_start_ts, drawdown_keywords)
    else:
        cash_start = 0.0
        loan_start = 0.0

    cash = cash_start
    loan_avail = loan_start

    for d in days:
        biz = d.date()
        a = daily_map.get(biz, {
            "actual_in_claim":0,"actual_in":0,"actual_out_oper":0,"principal":0,"interest":0,
            "drawdown":0,"loan_delta":0,"total_in":0,"total_out_cash":0,"cash_net":0
        })
        p = plan_map.get(biz, {"plan_in":0,"plan_out":0})

        cash_start_day = cash
        loan_start_day = loan_avail
        total_start = cash_start_day + loan_start_day

        cash_change = float(a["cash_net"]) + float(p["plan_in"]) - float(p["plan_out"])
        cash_end = cash_start_day + cash_change

        loan_end = loan_start_day + float(a["loan_delta"])
        total_end = cash_end + loan_end

        rows.append({
            "date": biz,
            "dow": DOW_KO[d.weekday()],
            "cash_start": int(round(cash_start_day, 0)),
            "loan_avail_start": int(round(loan_start_day, 0)),
            "total_start": int(round(total_start, 0)),

            "actual_in_claim": int(a["actual_in_claim"]),
            "actual_in": int(a["actual_in"]),
            "actual_out_oper": int(a["actual_out_oper"]),
            "principal": int(a["principal"]),
            "interest": int(a["interest"]),
            "drawdown": int(a["drawdown"]),
            "plan_in": int(p["plan_in"]),
            "plan_out": int(p["plan_out"]),

            "cash_change": int(round(cash_change, 0)),
            "loan_change": int(a["loan_delta"]),
            "total_change": int(round((cash_change + float(a["loan_delta"])), 0)),

            "cash_end": int(round(cash_end, 0)),
            "loan_avail_end": int(round(loan_end, 0)),
            "total_end": int(round(total_end, 0)),
        })

        cash = cash_end
        loan_avail = loan_end

    return pd.DataFrame(rows)

# =========================================================
# Plans DB (load/save)
# =========================================================
def _ensure_plan_cols(df: pd.DataFrame) -> pd.DataFrame:
    base_cols = ["date","direction","amount","label","month_key","updated_at"]
    if df is None or df.empty:
        return pd.DataFrame(columns=base_cols)
    for c in base_cols:
        if c not in df.columns:
            df[c] = ""
    return df[base_cols].copy()

def load_plans_month(year: int, month: int) -> pd.DataFrame:
    df = _ensure_plan_cols(read_df("plans")) if DB_ENABLED else _ensure_plan_cols(pd.DataFrame())
    mk = f"{year}-{month:02d}"
    if len(df) == 0:
        return _ensure_plan_cols(pd.DataFrame())
    out = df[df["month_key"].astype(str).eq(mk)].copy()

    # 타입 정리
    out["date"] = pd.to_datetime(out["date"], errors="coerce").dt.date.astype("object")
    out["direction"] = out["direction"].astype(str).replace({"입금":"IN","출금":"OUT"})
    out["amount"] = pd.to_numeric(out["amount"], errors="coerce").fillna(0).astype(int)
    out["label"] = out["label"].astype(str)
    out["month_key"] = mk
    return _ensure_plan_cols(out)

def save_plans_month(year: int, month: int, edited: pd.DataFrame):
    if not DB_ENABLED:
        return

    mk = f"{year}-{month:02d}"
    all_df = _ensure_plan_cols(read_df("plans"))
    # 기존 월 데이터 제거 후 덮어쓰기
    all_df = all_df[~all_df["month_key"].astype(str).eq(mk)].copy()

    e = edited.copy() if edited is not None else pd.DataFrame(columns=["date","direction","amount","label"])
    if len(e) > 0:
        e = e.copy()
        e["date"] = pd.to_datetime(e["date"], errors="coerce").dt.date.astype("object")
        e = e[e["date"].notna()].copy()
        e["direction"] = e["direction"].astype(str).str.upper().replace({"입금":"IN","출금":"OUT"})
        e["amount"] = pd.to_numeric(e["amount"], errors="coerce").fillna(0).astype(int)
        e["label"] = e.get("label","").astype(str)
        e["month_key"] = mk
        e["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        e = _ensure_plan_cols(e)
        all_df = pd.concat([all_df, e], ignore_index=True)

    overwrite_df("plans", all_df)

# =========================================================
# UI
# =========================================================
st.title("현금흐름 MVP (업로드 → 내부이동 자동검수(DB) → 월수입지출/가용금액(DB 예정))")

# ---- sidebar DB status + navigation
st.sidebar.header("DB 상태")
if DB_ENABLED:
    try:
        _ = gs_client()  # ensure auth
        st.sidebar.success("Google Sheets 연결 OK")
        it_df = read_df("internal_transfers")
        pl_df = read_df("plans")
        st.sidebar.write(f"• internal_transfers: {len(it_df):,} rows")
        st.sidebar.write(f"• plans: {len(pl_df):,} rows")
    except Exception as e:
        st.sidebar.error(f"Google Sheets 연결 실패: {e}")
        DB_ENABLED = False
else:
    st.sidebar.info("DB 미사용(Secrets 설정 필요)")

page = st.sidebar.radio("페이지", ["월수입지출(핵심)", "내부이동 검수", "일자별 요약", "원장(거래 목록)"])

st.sidebar.header("데이터 업로드")
files = st.sidebar.file_uploader(
    "엑셀 업로드 (여러개 가능) - 잔액정렬data.xlsx도 가능",
    type=["xlsx"],
    accept_multiple_files=True
)

st.sidebar.subheader("메디칼론 대출사용(인출) 키워드")
dd_kw_text = st.sidebar.text_input(
    "입금처(출금처)에 포함된 키워드(콤마로 구분)",
    value=",".join(DEFAULT_DRAWDOWN_KEYWORDS)
)
drawdown_keywords = [x.strip() for x in dd_kw_text.split(",") if x.strip()]

# ---- If no files, still render wireframe pages using empty tx
tx = pd.DataFrame(columns=[
    "tx_id","posted_at","biz_date","account_name","direction","subtype","amount",
    "counterparty","balance","is_excluded_account","is_internal_auto",
    "is_principal","is_interest","is_drawdown","source","file_name"
])
anchors = []
parse_errors = []

if files:
    all_tx = []
    for f in files:
        b = f.getvalue()
        try:
            tx_one, anchor = parse_any_excel(b)
            tx_one["file_name"] = f.name
            all_tx.append(tx_one)
            if anchor is not None:
                anchors.append(anchor)
        except Exception as e:
            parse_errors.append((f.name, str(e)))

    if parse_errors:
        st.error("일부 파일 파싱 실패")
        for fn, msg in parse_errors:
            st.write(f"- {fn}: {msg}")

    if all_tx:
        tx = pd.concat(all_tx, ignore_index=True).drop_duplicates(subset=["tx_id"]).sort_values("posted_at")

# ---- session state
if "confirmed_pairs" not in st.session_state:
    st.session_state.confirmed_pairs = set()

# --- internal_transfers(DB) load (1회)
if "it_status_map" not in st.session_state:
    st.session_state.it_status_map = {}
    if DB_ENABLED:
        try:
            it_df = load_it_db()
            st.session_state.it_status_map = it_map_from_df(it_df)
        except Exception:
            st.session_state.it_status_map = {}

# confirmed/rejected set 동기화
sync_confirmed_pairs_from_it_map()

# ---- anchor resolution
anchor_info = None
if anchors and not tx.empty:
    anchors_sorted = sorted(anchors, key=lambda a: a["loan_time"])
    a = anchors_sorted[-1]

    bank_cash_total, missing_accounts = compute_anchor_cash_from_bank_and_tx(tx, a)

    cash_time = a["cash_time"]
    loan_time = a["loan_time"]

    base = tx[~tx["is_excluded_account"]].copy()
    seg = base[(base["posted_at"] > cash_time) & (base["posted_at"] <= loan_time)].copy()
    seg["cash_delta"] = np.where(seg["direction"]=="IN", seg["amount"], -seg["amount"])
    cash_adjust = float(seg["cash_delta"].sum())
    anchor_cash_at_loan = float(bank_cash_total + cash_adjust)

    anchor_info = {
        "anchor_time": loan_time,
        "anchor_cash": anchor_cash_at_loan,
        "anchor_loan_avail": float(a["loan_avail"]),
        "missing_accounts": missing_accounts,
        "cash_time": cash_time,
        "loan_time": loan_time,
    }

# =========================================================
# Top: summary banner
# =========================================================
with st.container():
    left, right = st.columns([3,7])
    with left:
        if not tx.empty:
            min_dt = tx["posted_at"].min()
            max_dt = tx["posted_at"].max()
            st.metric("거래기간", f"{min_dt} ~ {max_dt}")
            st.metric("전체 거래(중복제거)", f"{len(tx):,}")
        else:
            st.info("아직 거래 파일이 업로드되지 않았어요. (예정 DB만으로 월말 예측 와이어프레임을 볼 수 있습니다.)")
    with right:
        if anchor_info is not None:
            cash = anchor_info["anchor_cash"]
            loan_av = anchor_info["anchor_loan_avail"]
            total = cash + loan_av
            c1, c2, c3 = st.columns(3)
            c1.metric("현금(병원 범위, 앵커)", fmt_int(cash))
            c2.metric("메디칼론 여유액(앵커)", fmt_int(loan_av))
            c3.metric("Total 가용금액(앵커)", fmt_int(total))
        else:
            st.caption("앵커(계좌별잔액+메디칼론한도여유액)를 포함한 파일을 업로드하면 가용금액 정합이 활성화됩니다.")

# =========================================================
# Page: Monthly core
# =========================================================
if page == "월수입지출(핵심)":
    st.header("월수입지출 (현금 + 메디칼론 + Total 가용금액)")

    # month selector
    today = date.today()
    if not tx.empty:
        months = sorted({(d.year, d.month) for d in pd.to_datetime(tx["biz_date"]).dropna().dt.to_pydatetime()})
        if not months:
            months = [(today.year, today.month)]
    else:
        months = [(today.year, today.month)]
    month_labels = [f"{y}-{m:02d}" for y,m in months]
    sel = st.selectbox("월 선택", month_labels, index=len(month_labels)-1)
    year, month = map(int, sel.split("-"))

    # manual anchor when no anchor_info
    if anchor_info is None:
        st.subheader("앵커 수동 입력(거래 업로드 전 와이어프레임용)")
        m1, m2, m3 = st.columns(3)
        with m1:
            manual_cash = st.number_input("월초 현금(병원 범위)", value=0, step=1000000)
        with m2:
            manual_loan = st.number_input("월초 메디칼론 여유액", value=0, step=1000000)
        with m3:
            st.metric("월초 Total 가용금액", fmt_int(manual_cash + manual_loan))
        # anchor_info를 “월초값 기준”으로 임시 구성(테이블 계산용)
        anchor_info_eff = {
            "anchor_time": pd.Timestamp(datetime.combine(date(year, month, 1), datetime.min.time())),
            "anchor_cash": float(manual_cash),
            "anchor_loan_avail": float(manual_loan),
            "missing_accounts": [],
            "cash_time": pd.Timestamp(datetime.combine(date(year, month, 1), datetime.min.time())),
            "loan_time": pd.Timestamp(datetime.combine(date(year, month, 1), datetime.min.time())),
        }
    else:
        anchor_info_eff = anchor_info

    # DB plans load
    st.subheader("예정 입력 (DB 저장/불러오기)")
    if "plan_df" not in st.session_state or st.session_state.get("plan_month_key") != f"{year}-{month:02d}":
        st.session_state.plan_df = load_plans_month(year, month) if DB_ENABLED else pd.DataFrame(columns=["date","direction","amount","label","month_key","updated_at"])
        st.session_state.plan_month_key = f"{year}-{month:02d}"

    # editor columns to show
    editable = st.session_state.plan_df.copy()
    # 편집에서는 month_key/updated_at 숨기고 핵심 4개만
    edit_view = editable[["date","direction","amount","label"]].copy()

    edit_view = st.data_editor(
        edit_view,
        use_container_width=True,
        num_rows="dynamic",
        column_config={
            "date": st.column_config.DateColumn("날짜"),
            "direction": st.column_config.SelectboxColumn("구분", options=["IN","OUT","입금","출금"]),
            "amount": st.column_config.NumberColumn("금액"),
            "label": st.column_config.TextColumn("항목/메모"),
        },
        key="plan_editor",
    )

    b_save, b_reload, b_export = st.columns([2,2,2])
    with b_save:
        if st.button("이번달 예정 DB 저장", use_container_width=True, disabled=(not DB_ENABLED)):
            save_plans_month(year, month, edit_view)
            st.success("저장 완료(구글시트 plans 반영).")
            st.session_state.plan_df = load_plans_month(year, month)
    with b_reload:
        if st.button("DB에서 다시 불러오기", use_container_width=True, disabled=(not DB_ENABLED)):
            st.session_state.plan_df = load_plans_month(year, month)
            st.success("불러오기 완료.")
    with b_export:
        st.download_button(
            "예정 CSV 다운로드",
            data=edit_view.to_csv(index=False).encode("utf-8-sig"),
            file_name=f"plans_{year}-{month:02d}.csv",
            mime="text/csv",
            use_container_width=True
        )

    # build month table
    table = build_month_table(
        tx=tx,
        plan_df=edit_view,
        year=year,
        month=month,
        anchor_info=anchor_info_eff,
        confirmed_pairs=st.session_state.confirmed_pairs,
        drawdown_keywords=drawdown_keywords
    )

    # headline projections
    last_row = table.iloc[-1] if len(table) else None
    if last_row is not None:
        p1, p2, p3 = st.columns(3)
        p1.metric("월말 예상 현금", fmt_int(last_row["cash_end"]))
        p2.metric("월말 예상 메디칼론 여유액", fmt_int(last_row["loan_avail_end"]))
        p3.metric("월말 예상 Total", fmt_int(last_row["total_end"]))

    st.subheader("일자별 롤링표 (가독성 중심)")
    disp = table.copy()
    disp["날짜"] = disp["date"].astype(str)
    disp["요일"] = disp["dow"]

    num_cols = [
        "total_start",
        "actual_in_claim","actual_in","actual_out_oper","principal","interest","drawdown",
        "plan_in","plan_out",
        "total_change",
        "total_end",
    ]
    # 표는 “핵심만” 먼저 보여주고, 상세는 expander로
    core_cols = ["날짜","요일"] + num_cols

    for c in core_cols:
        if c in disp.columns and c not in ["날짜","요일"]:
            disp[c] = disp[c].map(fmt_int)

    st.dataframe(disp[core_cols], use_container_width=True, hide_index=True)

    with st.expander("상세 컬럼 보기(현금/메디칼론 포함)"):
        disp2 = table.copy()
        disp2["날짜"] = disp2["date"].astype(str)
        disp2["요일"] = disp2["dow"]
        show_cols = [
            "날짜","요일",
            "cash_start","loan_avail_start","total_start",
            "actual_in_claim","actual_in","actual_out_oper","principal","interest","drawdown",
            "plan_in","plan_out",
            "cash_change","loan_change","total_change",
            "cash_end","loan_avail_end","total_end"
        ]
        for c in show_cols:
            if c not in ["날짜","요일"]:
                disp2[c] = disp2[c].map(fmt_int)
        st.dataframe(disp2[show_cols], use_container_width=True, hide_index=True)

    st.download_button(
        "월수입지출 CSV 다운로드",
        data=table.to_csv(index=False).encode("utf-8-sig"),
        file_name=f"monthly_available_{year}-{month:02d}.csv",
        mime="text/csv"
    )

    st.caption("※ 거래 파일을 업로드하면 actual(실제) 칼럼들이 자동으로 채워지고, 현재는 예정(plans) 기반 와이어프레임 모드로 동작합니다.")

# =========================================================
# Page: Internal transfer review
# =========================================================
if page == "내부이동 검수":
    st.header("내부이동 후보 검수 (자동 CONFIRMED + 예외만 수정 → DB 저장)")

    if tx.empty:
        st.info("거래 파일을 업로드하면 내부이동 후보가 생성됩니다.")
        st.stop()

    # DB에서 읽어온 상태
    if "it_status_map" not in st.session_state:
        st.session_state.it_status_map = {}
    it_map = st.session_state.it_status_map

    cand = build_internal_candidates(tx)
    if cand.empty:
        st.info("내부이동 후보가 없습니다.")
        st.stop()

    # ---------------------------------------------------------
    # 상단 컨트롤 (필터/일괄)
    # ---------------------------------------------------------
    cA, cB, cC, cD, cE = st.columns([2,2,2,2,4])

    with cA:
        show_rejected = st.checkbox("REJECTED 표시", value=False)
    with cB:
        only_unsaved = st.checkbox("미저장 변경만 보기", value=False)
    with cC:
        status_filter = st.multiselect(
            "상태 필터",
            options=["AUTO", "CONFIRMED", "REJECTED"],
            default=["CONFIRMED", "AUTO"],
        )
    with cD:
        save_clicked = st.button("DB 저장", use_container_width=True, disabled=(not DB_ENABLED))
    with cE:
        st.caption("기본은 신규 후보를 CONFIRMED로 제안합니다. 예외만 REJECTED/AUTO로 바꾸고 저장하세요.")

    # 후보 키 리스트
    cand["k_out"] = cand["out_tx_id"].astype(str)
    cand["k_in"] = cand["in_tx_id"].astype(str)
    keys = list(zip(cand["k_out"], cand["k_in"]))

    # ---------------------------------------------------------
    # 1) 기본값(신규 후보=CONFIRMED)을 세션에 미리 세팅
    #    - DB에 저장된 값이 있으면 그걸 우선
    # ---------------------------------------------------------
    for (out_id, in_id) in keys:
        k = (out_id, in_id)
        st_key = f"it_status_{out_id}_{in_id}"
        note_key = f"it_note_{out_id}_{in_id}"

        if st_key not in st.session_state:
            db_status = it_map.get(k, {}).get("status", "")
            if db_status in ["CONFIRMED", "REJECTED"]:
                st.session_state[st_key] = db_status
            else:
                st.session_state[st_key] = DEFAULT_NEW_CAND_STATUS  # 신규는 기본 CONFIRMED

        if note_key not in st.session_state:
            st.session_state[note_key] = it_map.get(k, {}).get("note", "")

    # ---------------------------------------------------------
    # 2) 일괄 버튼 (현재 후보 전체 CONFIRMED / AUTO)
    #    - 위에서 세션키를 미리 만들었으니 안정적으로 작동
    # ---------------------------------------------------------
    b1, b2, b3, b4 = st.columns([2,2,2,6])
    with b1:
        bulk_confirm = st.button("현재 후보 전체 CONFIRMED", use_container_width=True)
    with b2:
        bulk_auto = st.button("현재 후보 전체 AUTO(저장안함)", use_container_width=True)
    with b3:
        refresh_db = st.button("DB에서 새로고침", use_container_width=True, disabled=(not DB_ENABLED))
    with b4:
        st.caption("TIP: 대부분 내부이체면 ‘전체 CONFIRMED’ → 예외만 REJECTED로 바꾸고 저장이 가장 빠릅니다.")

    if refresh_db and DB_ENABLED:
        it_df = load_it_db()
        st.session_state.it_status_map = it_map_from_df(it_df)
        sync_confirmed_pairs_from_it_map()
        st.success("새로고침 완료")
        st.rerun()

    if bulk_confirm:
        for (out_id, in_id) in keys:
            st.session_state[f"it_status_{out_id}_{in_id}"] = "CONFIRMED"
            k = (out_id, in_id)
            st.session_state.it_status_map[k] = {"status": "CONFIRMED", "note": st.session_state.get(f"it_note_{out_id}_{in_id}", "")}
        sync_confirmed_pairs_from_it_map()
        st.rerun()

    if bulk_auto:
        for (out_id, in_id) in keys:
            st.session_state[f"it_status_{out_id}_{in_id}"] = "AUTO"
            k = (out_id, in_id)
            if k in st.session_state.it_status_map:
                st.session_state.it_status_map.pop(k, None)
        sync_confirmed_pairs_from_it_map()
        st.rerun()

    # ---------------------------------------------------------
    # 3) 필터 적용 (REJECTED 숨김/상태필터/미저장만)
    # ---------------------------------------------------------
    def effective_status(out_id: str, in_id: str) -> str:
        st_key = f"it_status_{out_id}_{in_id}"
        v = st.session_state.get(st_key, "AUTO")
        return v if v in ["AUTO","CONFIRMED","REJECTED"] else "AUTO"

    def effective_note(out_id: str, in_id: str) -> str:
        return st.session_state.get(f"it_note_{out_id}_{in_id}", "")

    def db_status_note(k):
        dbs = it_map.get(k, {}).get("status", "")
        dbn = it_map.get(k, {}).get("note", "")
        return dbs, dbn

    visible_mask = []
    for (out_id, in_id) in keys:
        k = (out_id, in_id)
        s_eff = effective_status(out_id, in_id)

        # 1) REJECTED 표시 옵션
        if (not show_rejected) and (s_eff == "REJECTED"):
            visible_mask.append(False)
            continue

        # 2) 상태 필터
        if s_eff not in status_filter:
            visible_mask.append(False)
            continue

        # 3) 미저장만 보기 (DB와 다르면 True)
        if only_unsaved:
            dbs, dbn = db_status_note(k)
            # AUTO는 저장 안 하므로 "미저장 변경" 판단을 이렇게:
            # - AUTO인데 DB에 CONFIRMED/REJECTED가 있으면 변경(=삭제효과) -> 저장 로직상 DB에서 제거는 안 하므로
            #   운영상 AUTO는 '저장대상 아님'으로 보고, 미저장 보기에서는 제외하는 게 혼란 적음
            if s_eff == "AUTO":
                visible_mask.append(False)
                continue
            if (dbs != s_eff) or (str(dbn) != str(effective_note(out_id, in_id))):
                visible_mask.append(True)
            else:
                visible_mask.append(False)
            continue

        visible_mask.append(True)

    cand_view = cand.loc[visible_mask].copy()

    # 요약 지표
    total_n = len(cand)
    visible_n = len(cand_view)
    confirmed_n = sum(1 for (o,i) in keys if effective_status(o,i) == "CONFIRMED")
    rejected_n = sum(1 for (o,i) in keys if effective_status(o,i) == "REJECTED")
    auto_n = sum(1 for (o,i) in keys if effective_status(o,i) == "AUTO")

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("후보(전체)", f"{total_n:,}")
    m2.metric("후보(표시중)", f"{visible_n:,}")
    m3.metric("CONFIRMED", f"{confirmed_n:,}")
    m4.metric("REJECTED / AUTO", f"{rejected_n:,} / {auto_n:,}")

    if cand_view.empty:
        st.info("필터 조건에 해당하는 후보가 없습니다.")
        st.stop()

    st.caption("AUTO=미결(저장 안함), CONFIRMED=내부이체 확정(집계 제외), REJECTED=내부이체 아님(다음부터 후보에서 숨김 가능)")

    # ---------------------------------------------------------
    # 4) 후보 리스트 렌더링 + 저장할 변경분 계산
    # ---------------------------------------------------------
    decisions_to_save = {}  # {(out,in): {'status':..., 'note':...}}

    for _, r in cand_view.sort_values("time_diff_seconds").iterrows():
        out_id = str(r["out_tx_id"]); in_id = str(r["in_tx_id"])
        k = (out_id, in_id)

        st_key = f"it_status_{out_id}_{in_id}"
        note_key = f"it_note_{out_id}_{in_id}"

        left, mid, right = st.columns([7,7,4])

        with left:
            st.markdown(f"**OUT** {r['out_time']} / {r['out_account']} / {fmt_int(r['amount'])}")
            if r.get("out_counterparty"):
                st.caption(f"{r['out_counterparty']}")

        with mid:
            st.markdown(f"**IN**  {r['in_time']} / {r['in_account']} / {fmt_int(r['amount'])}")
            if r.get("in_counterparty"):
                st.caption(f"{r['in_counterparty']}")
            st.caption(f"time diff: {int(r['time_diff_seconds'])}s")

        with right:
            sel = st.selectbox(
                "상태",
                options=IT_STATUSES,  # ["AUTO","CONFIRMED","REJECTED"]
                index=IT_STATUSES.index(st.session_state[st_key]) if st.session_state[st_key] in IT_STATUSES else 0,
                key=st_key,
            )
            note = st.text_input("메모", key=note_key)

            # 즉시 세션 map 반영
            if sel in ["CONFIRMED","REJECTED"]:
                st.session_state.it_status_map[k] = {"status": sel, "note": note}
            else:
                st.session_state.it_status_map.pop(k, None)

            # 저장 대상 판단(현재값 vs DB값)
            dbs = it_map.get(k, {}).get("status", "")
            dbn = it_map.get(k, {}).get("note", "")
            if sel in ["CONFIRMED","REJECTED"]:
                if (dbs != sel) or (str(dbn) != str(note)):
                    decisions_to_save[k] = {"status": sel, "note": note}

    # 집계 제외 동기화
    sync_confirmed_pairs_from_it_map()

    st.info(f"저장 예정 변경: {len(decisions_to_save):,}건 (DB 저장 버튼 누르면 반영됩니다.)")

    # ---------------------------------------------------------
    # 5) 저장 버튼 처리
    # ---------------------------------------------------------
    if save_clicked:
        upsert_it_db(decisions_to_save)
        # 저장 후 재로딩(정합)
        it_df = load_it_db()
        st.session_state.it_status_map = it_map_from_df(it_df)
        sync_confirmed_pairs_from_it_map()
        st.success(f"DB 저장 완료: {len(decisions_to_save):,}건")
        st.rerun()

# =========================================================
# Page: Daily summary
# =========================================================
elif page == "일자별 요약":
    st.header("일자별 요약")

    if tx.empty:
        st.info("거래 파일을 업로드하면 실제 입/출금 요약이 생성됩니다.")
    else:
        daily = daily_aggregations(tx, st.session_state.confirmed_pairs, drawdown_keywords).sort_values("biz_date")
        if daily.empty:
            st.info("데이터가 없습니다.")
        else:
            disp = daily.copy()
            disp["날짜"] = disp["biz_date"].astype(str)
            for col in ["actual_in_claim","actual_in","actual_out_oper","principal","interest","drawdown","loan_delta","cash_net"]:
                disp[col] = disp[col].map(fmt_int)
            st.dataframe(
                disp[["날짜","actual_in_claim","actual_in","actual_out_oper","principal","interest","drawdown","loan_delta","cash_net"]],
                use_container_width=True,
                hide_index=True
            )

# =========================================================
# Page: Ledger
# =========================================================
else:
    st.header("원장(거래 목록)")

    if tx.empty:
        st.info("거래 파일을 업로드하면 원장이 표시됩니다.")
    else:
        q = st.text_input("검색(계좌/상대/파일명)", value="")
        df = tx.copy()

        if q.strip():
            mask = (
                df["account_name"].astype(str).str.contains(q, na=False) |
                df["counterparty"].astype(str).str.contains(q, na=False) |
                df["file_name"].astype(str).str.contains(q, na=False)
            )
            df = df[mask]

        df = df.sort_values("posted_at", ascending=False).head(800)

        if drawdown_keywords:
            pat = "|".join([re.escape(k) for k in drawdown_keywords if k.strip()])
            is_dd = (df["direction"]=="IN") & (df["counterparty"].astype(str).str.contains(pat, na=False))
            is_dd = is_dd & (~df["counterparty"].astype(str).str.contains("원금|이자", na=False))
            df["is_drawdown_calc"] = is_dd
        else:
            df["is_drawdown_calc"] = False

        def loan_tag(r):
            if r["is_principal"]:
                return "원금상환"
            if r["is_drawdown_calc"]:
                return "대출사용"
            if r["is_interest"]:
                return "이자"
            return ""

        df["loan_event"] = df.apply(loan_tag, axis=1)

        disp = df[["posted_at","biz_date","account_name","direction","subtype","amount","counterparty","balance","loan_event","source","file_name"]].copy()
        disp["일시"] = disp["posted_at"].astype(str)
        disp["일자"] = disp["biz_date"].astype(str)
        disp["금액"] = disp["amount"].map(fmt_int)
        disp["잔액"] = disp["balance"].map(fmt_int)

        st.dataframe(
            disp[["일시","일자","account_name","direction","subtype","금액","counterparty","잔액","loan_event","source","file_name"]],
            use_container_width=True,
            hide_index=True
        )
